import openpyxl
import pandas as pd


def within_range(bounds: tuple, cell: openpyxl.cell) -> bool:
    """
    Функция, определяющая, является ли клетка слитой или нет
    :param bounds:
    :param cell:
    :return: True, если merged клетка, иначе False
    """
    column_start, row_start, column_end, row_end = bounds
    row = cell.row
    if row_start <= row <= row_end:
        column = cell.column
        if column_start <= column <= column_end:
            return True
    return False


def get_value_merged(sheet: openpyxl.worksheet, cell: openpyxl.cell) -> any:
    """
    Функция, возвращающая значение, лежащее в клетке, вне зависимости от того,
    является ли клетка merged, или нет
    :param sheet:
    :param cell:
    :return: значение, лежащее в клетке
    """
    for merged in sheet.merged_cells:
        if within_range(merged.bounds, cell):
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value


def get_timetable(table: openpyxl.worksheet) -> dict:
    """
    Функция, которая из таблицы Excel с расписанием выделяет расписание для каждой группы
    :param table: 
    :return: Словарь, ключи в котором являются номерами групп, а
    значение, соответствующее ключу – расписание для этой группы
    """
    groups = {}  # список расписаний для групп

    for j in range(3, table.max_column):  # смотрим на значения по столбцам
        name = table.cell(1, j).value  # номер группы
        if name in ['Дни', 'Часы']:  # если это не номер группы, то пропускаем столбец
            continue
        # иначе если столбец - это номер группы, то составляем для него расписание
        elif name is not None:
            # group - словарь с расписанием для группы
            group = dict(Понедельник={}, Вторник={}, Среда={}, Четверг={}, Пятница={}, Суббота={}, Воскресенье={})
            for k in range(2, table.max_row):  # проходимся по столбцу
                # если клетки относятся ко дню недели (не разделители)
                if get_value_merged(table, table.cell(k, 1)) in group.keys():
                    day = get_value_merged(table, table.cell(k, 1))  # значение дня недели
                    time = get_value_merged(table, table.cell(k, 2))  # клетка, в которой лежит значение времени
                    pair = get_value_merged(table, table.cell(k, j))  # клетка, в которой лежит значение пары

                    # рассматриваем только те клетки, для которых определено значение как пары, так и времени
                    if (time, pair) != (None, None):
                        time = time.split()  # преобразуем время пары к формату hh:mm – hh:mm
                        if len(time[0][:-2]) == 1:
                            time[0] = '0' + time[0]
                        time = time[0][:-2] + ':' + time[0][-2:] + ' – ' + time[2][:-2] + ':' + time[2][-2:]
                        group[day][time] = pair  # записываем значение в расписание

            group = pd.DataFrame(group)  # заменяем None на более красивые прочерки
            group.replace(to_replace=[None], value='–', inplace=True)
            groups[name] = group

    return groups


groups = []

kurs_1_doc = openpyxl.load_workbook('timetable/Raspisanie_1_kurs_osen_2019.xlsm')
kurs_1 = kurs_1_doc.active
groups.append(get_timetable(kurs_1))

kurs_2_doc = openpyxl.load_workbook('timetable/Raspisanie_2_kurs_osen_2019.xlsm')
kurs_2 = kurs_2_doc.active
groups.append(get_timetable(kurs_2))

kurs_3_doc = openpyxl.load_workbook('timetable/Raspisanie_3_kurs_osen_2019.xlsm')
kurs_3 = kurs_3_doc.active
groups.append(get_timetable(kurs_3))

kurs_4_doc = openpyxl.load_workbook('timetable/Raspisanie_4_kurs_osen_2019.xlsm')
kurs_4 = kurs_4_doc.active
groups.append(get_timetable(kurs_4))

kurs_5_doc = openpyxl.load_workbook('timetable/Raspisanie_5_kurs_osen_2019.xlsm')
kurs_5 = kurs_5_doc.active
groups.append(get_timetable(kurs_5))

kurs_6_faki_doc = openpyxl.load_workbook('timetable/Raspisanie_6_kurs_FAKI_osen_2019.xlsm')
kurs_6_faki = kurs_6_faki_doc.active
groups_6_faki = get_timetable(kurs_6_faki)

kurs_6_fupm_doc = openpyxl.load_workbook('timetable/Raspisanie_6_kurs_FUPM_osen_2019.xlsm')
kurs_6_fupm = kurs_6_fupm_doc.active
groups_6_fupm = get_timetable(kurs_6_fupm)

groups.append({**groups_6_faki, **groups_6_fupm})


def timetable_by_group(grade: int, group: str, day: str = 'week') -> pd.DataFrame or str:
    """
    Функция, выдающая расписание для нужной группы на требуемый день
    :param grade: номер курса
    :param group: номер группы
    :param day: день недели, расписание на который нужно вызвать, по умолчанию выдается расписание на неделю
    :return: расписание в формате pd.DataFrame
    """
    curr_groups = groups[grade - 1]  # смотрим, среди групп какого курса будем искать
    if group in curr_groups.keys() and day == 'week':  # если номер группы есть в списке, то выдаем нужное расписание
        return curr_groups[group]
    elif group in curr_groups.keys() and day != 'week':
        return curr_groups[group][day]
    else:  # иначе говорим, что произошла ошибка
        return []
